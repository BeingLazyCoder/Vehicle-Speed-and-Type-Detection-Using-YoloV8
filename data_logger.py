"""Logger for recording vehicle speed data in real-time"""
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import logging

logger = logging.getLogger(__name__)

# Speed limits per vehicle type (km/h)
SPEED_LIMITS = {
    'car': 120,
    'motorcycle': 120, 
    'bus': 80,
    'truck': 100,
    'van': 120
}

class SpeedLogger:
    def __init__(self):
        self.filename = "speed_violations.xlsx"
        self._initialize_excel()
        
    def _initialize_excel(self):
        """Initialize Excel file if it doesn't exist"""
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            headers = ["Timestamp", "Vehicle Type", "Speed", "Speed Limit", "Excess"]
            ws.append(headers)
            wb.save(self.filename)
        
    def log_violation(self, vehicle_type: str, speed: float):
        """Log speed violation immediately to Excel file"""
        vehicle_type = vehicle_type.lower()
        if vehicle_type not in SPEED_LIMITS:
            logger.warning(f"Unknown vehicle type: {vehicle_type}, using default limit")
            speed_limit = 150
        else:
            speed_limit = SPEED_LIMITS[vehicle_type]
        
        if speed > speed_limit:
            try:
                wb = load_workbook(self.filename)
                ws = wb.active
                
                violation_data = [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    vehicle_type.capitalize(),
                    f"{speed:.1f} km/h",
                    f"{speed_limit} km/h",
                    f"{speed - speed_limit:.1f} km/h"
                ]
                
                ws.append(violation_data)
                wb.save(self.filename)
                
            except Exception as e:
                logger.error(f"Error logging violation: {str(e)}")
    
    def save(self):
        """No need to save at the end as we're saving in real-time"""
        pass